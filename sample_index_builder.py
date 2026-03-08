#!/usr/bin/env python3
"""
sample_index_builder.py

Create and update master_sample_index.xlsx for quick sample lookup.
Columns: Batch ID | Sample ID | UCI ID | Date
"""

import os
import re
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import logging

from clinical_data_extractor import (
    normalize_sample_id,
    extract_clinical_from_order_sheet,
    find_order_sheet
)

logger = logging.getLogger(__name__)


def build_sample_index(
    batches: Dict[str, str],
    mutations_df: pd.DataFrame,
    output_path: str
) -> pd.DataFrame:
    """
    Build master sample index combining mutation samples with clinical data.

    Args:
        batches: Dict mapping batch_id to batch directory path
        mutations_df: DataFrame with consolidated mutations
        output_path: Path to save master_sample_index.xlsx

    Returns:
        DataFrame with sample index
    """
    # Get unique samples from mutations
    if mutations_df.empty or 'sample_ID' not in mutations_df.columns:
        logger.warning("No mutations data available for sample index")
        return pd.DataFrame()

    unique_samples = mutations_df[['batch_id', 'sample_ID']].drop_duplicates()

    # Add normalized sample ID for clinical linkage
    unique_samples['normalized_id'] = unique_samples['sample_ID'].apply(normalize_sample_id)

    # Build clinical data lookup
    clinical_lookup = {}

    for batch_id, batch_dir in batches.items():
        order_sheet = find_order_sheet(batch_dir)

        if order_sheet:
            clinical_df = extract_clinical_from_order_sheet(order_sheet)

            if not clinical_df.empty and 'D#' in clinical_df.columns:
                for _, row in clinical_df.iterrows():
                    key = (batch_id, str(row['D#']))
                    clinical_lookup[key] = {
                        'UCI_ID': row.get('UCI_ID', ''),
                        'Date': row.get('Date', ''),
                        'Age': row.get('Age', ''),
                        'Sex': row.get('Sex', ''),
                        'MPN': row.get('MPN', '')
                    }

    # Link samples to clinical data
    index_data = []

    for _, sample_row in unique_samples.iterrows():
        batch_id = sample_row['batch_id']
        sample_id = sample_row['sample_ID']
        normalized_id = sample_row['normalized_id']

        # Try to find clinical data
        clinical = clinical_lookup.get((batch_id, normalized_id), {})

        index_data.append({
            'Batch_ID': batch_id,
            'Sample_ID': sample_id,
            'UCI_ID': clinical.get('UCI_ID', ''),
            'Date': clinical.get('Date', ''),
            'Age': clinical.get('Age', ''),
            'Sex': clinical.get('Sex', ''),
            'MPN': clinical.get('MPN', '')
        })

    index_df = pd.DataFrame(index_data)

    # Sort by Batch_ID and Sample_ID
    index_df = index_df.sort_values(['Batch_ID', 'Sample_ID'])

    # Save to Excel with formatting
    save_sample_index(index_df, output_path)

    return index_df


def save_sample_index(df: pd.DataFrame, output_path: str) -> None:
    """
    Save sample index to Excel with proper formatting.

    Args:
        df: Sample index DataFrame
        output_path: Path to save Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active
        ws.title = "Sample Index"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                if r_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        # Adjust column widths
        column_widths = {
            'A': 18,  # Batch_ID
            'B': 15,  # Sample_ID
            'C': 15,  # UCI_ID
            'D': 12,  # Date
            'E': 8,   # Age
            'F': 6,   # Sex
            'G': 10   # MPN
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze top row
        ws.freeze_panes = "A2"

        # Add filter
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        logger.info(f"Saved sample index to {output_path}")

    except ImportError:
        # Fallback to basic Excel save
        df.to_excel(output_path, index=False)
        logger.info(f"Saved sample index (basic format) to {output_path}")


def update_sample_index(
    new_batches: Dict[str, str],
    new_mutations: pd.DataFrame,
    existing_index_path: str
) -> pd.DataFrame:
    """
    Update existing sample index with new batches.

    Args:
        new_batches: Dict of new batch_id to directory mappings
        new_mutations: DataFrame with mutations from new batches
        existing_index_path: Path to existing master_sample_index.xlsx

    Returns:
        Updated sample index DataFrame
    """
    # Load existing index
    if os.path.exists(existing_index_path):
        existing_df = pd.read_excel(existing_index_path)
    else:
        existing_df = pd.DataFrame()

    # Build index for new batches
    new_index = build_sample_index(
        new_batches,
        new_mutations,
        existing_index_path + '.temp'
    )

    # Remove temp file if created
    temp_path = existing_index_path + '.temp'
    if os.path.exists(temp_path):
        os.remove(temp_path)

    # Combine existing and new
    if not existing_df.empty and not new_index.empty:
        combined = pd.concat([existing_df, new_index], ignore_index=True)
        combined = combined.drop_duplicates(subset=['Batch_ID', 'Sample_ID'])
    elif not new_index.empty:
        combined = new_index
    else:
        combined = existing_df

    # Save combined index
    save_sample_index(combined, existing_index_path)

    return combined


def search_sample(
    index_df: pd.DataFrame,
    query: str,
    column: str = None
) -> pd.DataFrame:
    """
    Search for samples in the index.

    Args:
        index_df: Sample index DataFrame
        query: Search query (partial match supported)
        column: Specific column to search, or None for all

    Returns:
        DataFrame with matching samples
    """
    if index_df.empty:
        return pd.DataFrame()

    query_lower = query.lower()

    if column and column in index_df.columns:
        mask = index_df[column].astype(str).str.lower().str.contains(query_lower, na=False)
    else:
        # Search across all text columns
        mask = pd.Series([False] * len(index_df))
        for col in index_df.columns:
            mask |= index_df[col].astype(str).str.lower().str.contains(query_lower, na=False)

    return index_df[mask]


def get_sample_summary(index_df: pd.DataFrame) -> Dict:
    """
    Generate summary statistics for the sample index.

    Args:
        index_df: Sample index DataFrame

    Returns:
        Dict with summary statistics
    """
    if index_df.empty:
        return {}

    summary = {
        'total_samples': len(index_df),
        'total_batches': index_df['Batch_ID'].nunique(),
        'batches': index_df['Batch_ID'].value_counts().to_dict()
    }

    # Sex distribution
    if 'Sex' in index_df.columns:
        summary['sex_distribution'] = index_df['Sex'].value_counts().to_dict()

    # MPN distribution
    if 'MPN' in index_df.columns:
        summary['mpn_distribution'] = index_df['MPN'].value_counts().to_dict()

    # Age statistics
    if 'Age' in index_df.columns:
        ages = pd.to_numeric(index_df['Age'], errors='coerce').dropna()
        if len(ages) > 0:
            summary['age_stats'] = {
                'mean': ages.mean(),
                'median': ages.median(),
                'min': ages.min(),
                'max': ages.max()
            }

    return summary


if __name__ == '__main__':
    import sys

    logging.basicConfig(level=logging.INFO)

    # Test with sample data
    test_mutations = pd.DataFrame({
        'batch_id': ['KG001_01.22.25', 'KG001_01.22.25', '10_23_25'],
        'sample_ID': ['D98A,D98B', 'D99A,D99B', 'D100A']
    })

    smmip_root = os.environ.get("SMMIP_ROOT", ".")
    test_batches = {
        'KG001_01.22.25': os.path.join(smmip_root, 'KG001_01.22.25'),
        '10_23_25': os.path.join(smmip_root, '10_23_25'),
    }

    # Check which batches exist
    existing_batches = {k: v for k, v in test_batches.items() if os.path.isdir(v)}

    if existing_batches:
        print(f"Testing with batches: {list(existing_batches.keys())}")

        output = os.path.join(smmip_root, 'Master_Output', 'test_sample_index.xlsx')
        df = build_sample_index(existing_batches, test_mutations, output)

        print(f"\nSample index ({len(df)} rows):")
        print(df)

        print("\nSummary:")
        print(get_sample_summary(df))

        # Cleanup
        if os.path.exists(output):
            os.remove(output)
    else:
        print("No test batches found")
